from openpyxl import load_workbook
import json

# Load Excel file using openpyxl
workbook = load_workbook('sample_data.xlsx')
sheet = workbook.active

print("=== EXCEL FILE ANALYSIS ===")
print(f"Sheet name: {sheet.title}")
print(f"Max row: {sheet.max_row}")
print(f"Max column: {sheet.max_column}")

# Get headers
headers = []
for col in range(1, sheet.max_column + 1):
    cell_value = sheet.cell(row=1, column=col).value
    headers.append(cell_value)

print(f"\nHeaders: {headers}")

# Get sample data
print("\n=== SAMPLE DATA ===")
sample_data = []
for row in range(1, min(6, sheet.max_row + 1)):  # First 5 rows
    row_data = []
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=row, column=col).value
        row_data.append(cell_value)
    sample_data.append(row_data)
    print(f"Row {row}: {row_data}")

# Analyze each column
print("\n=== COLUMN ANALYSIS ===")
for col_idx, header in enumerate(headers, 1):
    print(f"\nColumn: {header}")
    
    # Get all values for this column (excluding header)
    values = []
    for row in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=col_idx).value
        if cell_value is not None:
            values.append(cell_value)
    
    print(f"  Total non-null values: {len(values)}")
    print(f"  Unique values: {len(set(values))}")
    
    # Sample values
    sample_values = values[:5] if values else []
    print(f"  Sample values: {sample_values}")
    
    # Data type analysis
    if values:
        value_types = set(type(v).__name__ for v in values)
        print(f"  Data types: {value_types}")
        
        # String length analysis for text data
        if any(isinstance(v, str) for v in values):
            str_values = [v for v in values if isinstance(v, str)]
            if str_values:
                max_len = max(len(v) for v in str_values)
                avg_len = sum(len(v) for v in str_values) / len(str_values)
                print(f"  String lengths - Max: {max_len}, Avg: {avg_len:.1f}")

# Save analyzed data to JSON for further processing
data_dict = {
    'headers': headers,
    'total_rows': sheet.max_row - 1,  # Excluding header
    'total_columns': sheet.max_column,
    'sample_data': sample_data[1:6]  # Excluding header row
}

with open('data_analysis.json', 'w', encoding='utf-8') as f:
    json.dump(data_dict, f, ensure_ascii=False, indent=2, default=str)

print(f"\n=== SUMMARY ===")
print(f"Total data rows: {sheet.max_row - 1}")
print(f"Total columns: {sheet.max_column}")
print(f"Data saved to: data_analysis.json")